"""Tests for the Heimspielplan xlsx parser."""

import io
from datetime import date, time

import openpyxl
import pytest

from app.models.planning import SeasonType, SpielTyp
from app.services.game_plan_parser import GamePlanParseError, parse_game_plan

HEADERS = [
    "SpielTyp",
    "Bezeichnung",
    "Spielnummer",
    "TagKurz",
    "Spieldatum",
    "Spielzeit",
    "Teams",
    "Spielort",
    "Bemerkung",
]


def _workbook(sheets: dict[str, list[list[object]]]) -> bytes:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets.items():
        worksheet = workbook.create_sheet(title)
        worksheet.append(HEADERS)
        for row in rows:
            worksheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parses_hr_and_rr_sheets_into_autumn_and_spring() -> None:
    content = _workbook(
        {
            "HR 2026 2027": [
                [
                    "Meisterschaft",
                    "FCTC Herren 1/4. Liga",
                    119550,
                    "Sa",
                    date(2026, 8, 15),
                    time(17, 0),
                    "FC Thusis/Cazis 1 - FC Lusitanos",
                    "St. Martin, Cazis - Platz 1",
                    None,
                ]
            ],
            "RR 2025 2026": [
                [
                    "Cup",
                    "Cup Spiel",
                    515539,
                    "Sa",
                    date(2026, 4, 4),
                    time(17, 0),
                    "FC Thusis/Cazis 1 - FC Rapperswil",
                    "St. Martin, Cazis - Platz 1",
                    None,
                ]
            ],
        }
    )
    sheets = parse_game_plan(content)
    by_tab = {sheet.sheet_tab: sheet for sheet in sheets}
    assert by_tab["HR 2026 2027"].season_type == SeasonType.AUTUMN
    assert by_tab["RR 2025 2026"].season_type == SeasonType.SPRING
    assert by_tab["HR 2026 2027"].rows[0].spiel_typ == SpielTyp.MEISTERSCHAFT
    assert by_tab["HR 2026 2027"].rows[0].spielnummer == "119550"


def test_spielnummer_ohne_becomes_none() -> None:
    content = _workbook(
        {
            "HR 2026 2027": [
                [
                    "Testspiel",
                    "Testspiel FCTC",
                    "Ohne",
                    "Mi",
                    date(2026, 8, 5),
                    time(20, 0),
                    "FCTC Grp - FCTC 2",
                    "St. Martin, Cazis - Platz 1",
                    "  ",
                ]
            ]
        }
    )
    rows = parse_game_plan(content)[0].rows
    assert rows[0].spielnummer is None
    assert rows[0].remark is None


def test_trainingsspiele_maps_to_testspiel() -> None:
    content = _workbook(
        {
            "RR 2025 2026": [
                [
                    "Trainingsspiele",
                    "Trainingsspiel",
                    "Ohne",
                    "Sa",
                    date(2026, 3, 21),
                    time(16, 30),
                    "FC Thusis/Cazis 1 - FC Landquart",
                    "Obere Au, Chur",
                    None,
                ]
            ]
        }
    )
    rows = parse_game_plan(content)[0].rows
    assert rows[0].spiel_typ == SpielTyp.TESTSPIEL


def test_blank_row_without_date_is_skipped() -> None:
    content = _workbook(
        {
            "HR 2026 2027": [
                [None, None, None, None, None, None, None, None, None],
                [
                    "Cup",
                    "Cup Spiel",
                    "Ohne",
                    "Sa",
                    date(2026, 8, 8),
                    time(17, 0),
                    "A - B",
                    "St. Martin, Cazis",
                    None,
                ],
            ]
        }
    )
    rows = parse_game_plan(content)[0].rows
    assert len(rows) == 1


def test_unknown_sheet_prefix_is_ignored() -> None:
    content = _workbook(
        {
            "Sonstiges": [
                [
                    "Cup",
                    "Cup Spiel",
                    "Ohne",
                    "Sa",
                    date(2026, 8, 8),
                    time(17, 0),
                    "A - B",
                    "St. Martin, Cazis",
                    None,
                ]
            ]
        }
    )
    with pytest.raises(GamePlanParseError):
        parse_game_plan(content)


def test_unknown_spiel_typ_raises() -> None:
    content = _workbook(
        {
            "HR 2026 2027": [
                [
                    "Freundschaftsspiel",
                    "Unbekannt",
                    "Ohne",
                    "Sa",
                    date(2026, 8, 8),
                    time(17, 0),
                    "A - B",
                    "St. Martin, Cazis",
                    None,
                ]
            ]
        }
    )
    with pytest.raises(GamePlanParseError):
        parse_game_plan(content)


def test_missing_header_raises() -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "HR 2026 2027"
    worksheet.append(["Datum", "Zeit"])
    worksheet.append([date(2026, 8, 8), time(17, 0)])
    buffer = io.BytesIO()
    workbook.save(buffer)
    with pytest.raises(GamePlanParseError):
        parse_game_plan(buffer.getvalue())


def test_missing_teams_or_venue_raises() -> None:
    content = _workbook(
        {
            "HR 2026 2027": [
                [
                    "Cup",
                    "Cup Spiel",
                    "Ohne",
                    "Sa",
                    date(2026, 8, 8),
                    time(17, 0),
                    "",
                    "St. Martin, Cazis",
                    None,
                ]
            ]
        }
    )
    with pytest.raises(GamePlanParseError):
        parse_game_plan(content)
