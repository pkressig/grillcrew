"""Pure parser for the coordinator's matrix-style Kiosk/Grill workbook."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from enum import StrEnum

from openpyxl import load_workbook


class ExternalKioskParseError(ValueError):
    pass


class ExternalKioskCategory(StrEnum):
    KIOSK = "KIOSK"
    GRILL = "GRILL"


@dataclass(frozen=True)
class ParsedExternalKioskRow:
    sheet_name: str
    source_row: int
    plan_date: date
    start_time: time | None
    end_time: time | None
    category: ExternalKioskCategory
    assigned_person_text: str | None
    source_note: str | None
    raw_date: str
    raw_time: str


_RANGE = re.compile(
    r"(?P<start>\d{1,2}[.:]\d{2})\s*-\s*(?P<end>\d{1,2}[.:]\d{2}|Schluss|Ende)", re.I
)
_EMPTY = {"", "--", "-", "kein", "nein"}


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value).strip()


def _raw_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value)


def _date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    match = re.search(r"(\d{1,2})[./](?:\d{1,2}[./])?(\d{1,2})[./](\d{4})", text)
    if match:
        # For combined values such as 03./04.09.2025, retain the first date.
        day = int(re.match(r"\d{1,2}", text).group())  # type: ignore[union-attr]
        return date(int(match.group(3)), int(match.group(2)), day)
    return None


def _clock(value: str) -> time:
    hour, minute = (int(part) for part in value.replace(":", ".").split("."))
    return time(hour, minute)


def _ranges(raw: str, fallback_end: time | None = None) -> list[tuple[time, time | None]]:
    result: list[tuple[time, time | None]] = []
    for match in _RANGE.finditer(raw):
        end_text = match.group("end").lower()
        end = fallback_end if end_text in {"schluss", "ende"} else _clock(end_text)
        result.append((_clock(match.group("start")), end))
    return result


def _meaningful(value: object) -> bool:
    return _text(value).lower().strip() not in _EMPTY


def parse_external_kiosk_plan(content: bytes) -> list[ParsedExternalKioskRow]:
    """Parse workbook bytes without database or filesystem side effects."""
    if not content:
        raise ExternalKioskParseError("Die Excel-Datei ist leer.")
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise ExternalKioskParseError("Die Datei ist keine lesbare Excel-Arbeitsmappe.") from error
    if "Tabelle1" not in workbook.sheetnames:
        raise ExternalKioskParseError("Das erwartete Tabellenblatt 'Tabelle1' fehlt.")
    sheet = workbook["Tabelle1"]
    headers = [_text(cell.value) for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    if len(headers) < 13 or headers[0] != "Datum" or headers[1] != "Zeiten":
        raise ExternalKioskParseError("Die Kiosk-Matrix hat nicht die erwarteten Spalten.")
    people = headers[2:11]
    parsed: list[ParsedExternalKioskRow] = []
    for source_row, cells in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        values = list(cells) + [None] * max(0, 15 - len(cells))
        kiosk_date = _date(values[0])
        kiosk_raw = _text(values[1])
        kiosk_ranges = _ranges(kiosk_raw)
        if kiosk_date is not None and _meaningful(values[1]):
            start, end = kiosk_ranges[0] if kiosk_ranges else (None, None)
            parsed.append(
                ParsedExternalKioskRow(
                    sheet.title,
                    source_row,
                    kiosk_date,
                    start,
                    end,
                    ExternalKioskCategory.KIOSK,
                    None,
                    _raw_text(values[14]) or None,
                    _raw_text(values[0]),
                    _raw_text(values[1]),
                )
            )
            for name, value in zip(people, values[2:11], strict=True):
                if not name or not _meaningful(value):
                    continue
                assignment_raw = _raw_text(value)
                assignment_ranges = _ranges(assignment_raw.strip(), end)
                assignment_start, assignment_end = (
                    assignment_ranges[0] if assignment_ranges else (None, None)
                )
                parsed.append(
                    ParsedExternalKioskRow(
                        sheet.title,
                        source_row,
                        kiosk_date,
                        assignment_start,
                        assignment_end,
                        ExternalKioskCategory.KIOSK,
                        name,
                        assignment_raw,
                        _raw_text(values[0]),
                        assignment_raw,
                    )
                )

        grill_date = _date(values[11])
        grill_raw = _text(values[12])
        if grill_date is not None and _meaningful(values[12]):
            grill_ranges = _ranges(grill_raw)
            note_parts = [part for part in (_raw_text(values[13]), _raw_text(values[14])) if part]
            normalized_ranges: list[tuple[time | None, time | None]] = list(grill_ranges)
            if not normalized_ranges:
                normalized_ranges.append((None, None))
            for start, end in normalized_ranges:
                parsed.append(
                    ParsedExternalKioskRow(
                        sheet.title,
                        source_row,
                        grill_date,
                        start,
                        end,
                        ExternalKioskCategory.GRILL,
                        None,
                        "; ".join(note_parts) or None,
                        _raw_text(values[11]),
                        _raw_text(values[12]),
                    )
                )
    if not parsed:
        raise ExternalKioskParseError("Die Arbeitsmappe enthält keine Kiosk- oder Grillzeilen.")
    return parsed
