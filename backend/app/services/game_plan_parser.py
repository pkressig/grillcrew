"""Parse the football association's Heimspielplan export into structured rows.

Pure, side-effect-free parsing: reads workbook bytes and returns dataclasses only. Season/
club-year resolution, home-venue matching, and diffing against existing events happen in
``app.services.imports``.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date as date_type
from datetime import datetime
from datetime import time as time_type
from io import BytesIO

from openpyxl import load_workbook

from app.models.planning import SeasonType, SpielTyp

SPIEL_TYP_MAP: dict[str, SpielTyp] = {
    "meisterschaft": SpielTyp.MEISTERSCHAFT,
    "cup": SpielTyp.CUP,
    "testspiel": SpielTyp.TESTSPIEL,
    "trainingsspiele": SpielTyp.TESTSPIEL,
    "turnier": SpielTyp.TURNIER,
}

SHEET_ROUND_PATTERN = re.compile(r"^\s*(HR|RR)\b", re.IGNORECASE)
SHEET_ROUND_SEASON_TYPE: dict[str, SeasonType] = {"HR": SeasonType.AUTUMN, "RR": SeasonType.SPRING}

REQUIRED_HEADERS: tuple[str, ...] = (
    "SpielTyp",
    "Bezeichnung",
    "Spielnummer",
    "TagKurz",
    "Spieldatum",
    "Spielzeit",
    "Teams",
    "Spielort",
    "Bemerkung",
)
_HEADER_SCAN_ROWS = 15


class GamePlanParseError(Exception):
    pass


@dataclass(frozen=True)
class ParsedRow:
    sheet_tab: str
    spiel_typ: SpielTyp
    bezeichnung: str | None
    spielnummer: str | None
    weekday_short: str | None
    game_date: date_type
    game_time: time_type | None
    teams: str
    venue: str
    remark: str | None


@dataclass(frozen=True)
class ParsedSheet:
    sheet_tab: str
    season_type: SeasonType
    rows: list[ParsedRow]


def parse_game_plan(content: bytes) -> list[ParsedSheet]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as error:
        raise GamePlanParseError("die Datei konnte nicht als Excel-Datei gelesen werden") from error

    sheets: list[ParsedSheet] = []
    for worksheet in workbook.worksheets:
        match = SHEET_ROUND_PATTERN.match(worksheet.title)
        if not match:
            continue
        season_type = SHEET_ROUND_SEASON_TYPE[match.group(1).upper()]
        header_row_index, columns = _find_header_row(worksheet)
        rows = [
            row
            for row in (
                _parse_row(worksheet.title, columns, raw_row)
                for raw_row in worksheet.iter_rows(min_row=header_row_index + 1)
            )
            if row is not None
        ]
        sheets.append(ParsedSheet(sheet_tab=worksheet.title, season_type=season_type, rows=rows))

    if not sheets:
        raise GamePlanParseError("keine Tabellenblätter mit einem 'HR'- oder 'RR'-Namen gefunden")
    return sheets


def _find_header_row(worksheet: object) -> tuple[int, dict[str, int]]:
    rows = getattr(worksheet, "iter_rows")(max_row=_HEADER_SCAN_ROWS)  # noqa: B009
    for row_index, row in enumerate(rows, start=1):
        values: dict[str, int] = {}
        for cell in row:
            cell_value = getattr(cell, "value", None)
            if cell_value is None:
                continue
            values[str(cell_value).strip()] = getattr(cell, "column") - 1  # noqa: B009
        if all(header in values for header in REQUIRED_HEADERS):
            return row_index, {header: values[header] for header in REQUIRED_HEADERS}
    raise GamePlanParseError(
        f"Kopfzeile mit den Spalten {', '.join(REQUIRED_HEADERS)} wurde nicht gefunden"
    )


def _parse_row(sheet_tab: str, columns: dict[str, int], row: object) -> ParsedRow | None:
    cells = list(row)  # type: ignore[call-overload]

    def value(header: str) -> object:
        index = columns[header]
        if index >= len(cells):
            return None
        return getattr(cells[index], "value", None)

    raw_date = value("Spieldatum")
    if raw_date is None:
        return None

    spiel_typ_raw = str(value("SpielTyp") or "").strip()
    spiel_typ = SPIEL_TYP_MAP.get(spiel_typ_raw.casefold())
    if spiel_typ is None:
        raise GamePlanParseError(f"unbekannter SpielTyp '{spiel_typ_raw}' in Blatt '{sheet_tab}'")

    game_date = raw_date.date() if isinstance(raw_date, datetime) else raw_date
    if not isinstance(game_date, date_type):
        raise GamePlanParseError(f"ungültiges Spieldatum in Blatt '{sheet_tab}'")

    raw_time = value("Spielzeit")
    game_time = raw_time if isinstance(raw_time, time_type) else None

    raw_number = value("Spielnummer")
    spielnummer: str | None = None
    if raw_number is not None:
        text_number = str(raw_number).strip()
        if text_number and text_number.casefold() != "ohne":
            spielnummer = text_number

    teams = str(value("Teams") or "").strip()
    venue = str(value("Spielort") or "").strip()
    if not teams or not venue:
        raise GamePlanParseError(f"Teams oder Spielort fehlen in Blatt '{sheet_tab}'")

    bezeichnung = str(value("Bezeichnung") or "").strip() or None
    weekday_short = str(value("TagKurz") or "").strip() or None
    raw_remark = value("Bemerkung")
    remark = str(raw_remark).strip() or None if raw_remark is not None else None

    return ParsedRow(
        sheet_tab=sheet_tab,
        spiel_typ=spiel_typ,
        bezeichnung=bezeichnung,
        spielnummer=spielnummer,
        weekday_short=weekday_short,
        game_date=game_date,
        game_time=game_time,
        teams=teams,
        venue=venue,
        remark=remark,
    )
